import pandas as pd
from cellStyleEnum import CellStyle
from copy import copy
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font


class ConstrutorOutput:
    def __init__(self, df: pd.DataFrame):
        self.df = df
        self.wb_data = Workbook()
        self.ws_data = self.wb_data.active

        # Carregando dados necessárias para o workbook
        self.__carrega_fomatacao_wb(wb=self.wb_data)
        self.__popular_wb_dados()

        self.__formatar_wb_dados()

    @staticmethod
    def __carrega_fomatacao_wb(wb: Workbook):
        """
        Carrega formatações em um dado workbook.
        """
        for style in CellStyle:
            wb.add_named_style(style.value)

    def __popular_wb_dados(self):
        """
        Função responsável por popular o workbook de dados com os presentes no dataframe fornecido no
        construtor.
        """
        for r_idx, row in enumerate(dataframe_to_rows(self.df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                self.ws_data.cell(row=r_idx, column=c_idx, value=value)

    def __formatar_wb_dados(self):
        """
        Aplica as formatações necessárias para o workbook de output
        """
        # Formatando celulas na coluna Resultados
        col_idx = self.df.columns.get_loc("Resultado") + 1
        for col in self.ws_data.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in col:
                if "<" in cell.value:
                    cell.style = CellStyle.CINZA_BG_STYLE.value   # fundo cinza caso resultado < LQ
                else:
                    existing_color = cell.font.color if cell.font.color else None   # Negrito caso resultado >= LQ
                    cell.font = Font(bold=True, color=existing_color)

        # Centraliza celulas no workbook
        for col in self.ws_data.iter_cols(min_col=1, max_col=self.ws_data.max_column):
            for r_idx, cell in enumerate(col, 1):
                alignment_obj = copy(cell.alignment)
                alignment_obj.horizontal = 'center'
                alignment_obj.vertical = 'center'
                cell.alignment = alignment_obj

    def salvar_workbook_formatado(self, output_path: str):
        """
        Salva o workbook formatado no local desejado

        :param output_path: Path para criação do arquivo de excel
        :type output_path: str
        """
        self.wb_data.save(output_path)
