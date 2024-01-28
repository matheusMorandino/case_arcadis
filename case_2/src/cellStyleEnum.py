from enum import Enum
from openpyxl.styles import NamedStyle, PatternFill, Font


class CellStyle(Enum):
    CINZA_BG_STYLE = NamedStyle(name='gray_bg', fill=PatternFill(start_color='B0B0B0', end_color='B0B0B0', fill_type='solid'))
    LARANJA_FONT_STYLE = NamedStyle(name='orange_text', font=Font(color="FFA500"))
    CINZA_BG_LARANJA_FONT_STYLE = NamedStyle(name='gray_bg_orange_text', font=Font(color="FFA500"), fill=PatternFill(start_color='B0B0B0', end_color='B0B0B0', fill_type='solid'))
