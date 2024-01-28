import os
import pandas as pd
from cellStyleEnum import CellStyle
from copy import copy
from itertools import product
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle, PatternFill, Font, Border, Alignment

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


class OutputConstrutor:
    def __init__(self, df: pd.DataFrame):
        self.df = df
        self.wb_header = self.__instancia_wb_header()
        self.ws_header = self.__instancia_ws_header()
        self.max_linha_header = self.ws_header.max_row
        self.wb_data = Workbook()
        self.ws_data = self.wb_data.active

        # Carregando dados necessárias para o workbook
        self.__carrega_fomatacao_wb(wb=self.wb_data)
        self.__popular_wb_dados()

        self.__formatar_wb_dados()
        self.__remover_elementos_desncessarios_wb_dados()

        self.__unir_conteudo_wb_header_dados()

        self.__mesclar_células_nan()

    @staticmethod
    def __is_nan(num):
        """
        Retorna se um valor é nan ou não.

        :param num:
        :return: boolean
        """
        return num != num

    @staticmethod
    def __carrega_fomatacao_wb(wb: Workbook):
        """
        Carrega formatações em um dado workbook.
        """
        for style in CellStyle:
            wb.add_named_style(style.value)


    @staticmethod
    def __instancia_wb_header():
        """
        Carrega o workbook que contem o cabeçalho necessário para o output

        :return: openpyxl.Workbook
        """
        internal_folder_path = os.path.join(ROOT_DIR, "_internal")

        if os.path.exists(internal_folder_path) and os.path.isdir(internal_folder_path):
            existing_file_path = os.path.join(ROOT_DIR, "_internal", "output_headers.xlsx")
        else:
            existing_file_path = os.path.join(ROOT_DIR, "output_headers.xlsx")

        wb_header = load_workbook(existing_file_path)

        return wb_header

    def merge_cells_below_nan(self, col: tuple, start_row: int):
        """
        Mescla todas as células com nan com a ultima célula não nan

        :param col: tuple com as células da coluna
        :type col: tuple
        :param start_row: linha inicial dos dados
        :type start_row: int
        :return:
        """
        nan_period = False
        inicio_cell = f'{col[start_row].column_letter}{1}'

        for r_idx, cell in enumerate(col, 1):
            if not self.__is_nan(cell.value):
                if not nan_period:
                    inicio_cell = f'{col[start_row].column_letter}{r_idx}'
                else:
                    fim_cell = f'{col[start_row].column_letter}{r_idx - 1}'
                    self.ws_header.merge_cells(f'{inicio_cell}:{fim_cell}')
                    inicio_cell = f'{col[start_row].column_letter}{r_idx}'
                    nan_period = False

            if self.__is_nan(cell.value):
                if not nan_period:
                    nan_period = True

        # Lidando com o caso especial da linha final
        if nan_period:
            fim_cell = f'{col[start_row].column_letter}{len(col)}'
            self.ws_header.merge_cells(f'{inicio_cell}:{fim_cell}')

    def __instancia_ws_header(self):
        """
        Retorna uma instância de worksheet para o cabeçalho do output, além de carregas as céluas
        no worksheet

        :return: openpyxl.Worksheet
        """
        ws_header = self.wb_header.active

        merged_cells_info = {}
        for merged_cells_range in ws_header.merged_cells.ranges:
            start_cell = merged_cells_range.coord.split(':')[0]
            merged_cells_info[start_cell] = merged_cells_range.size

        return ws_header

    def style_cells_cma(self, alvo_cols: list, flag_cols: list):
        """
        Realiza a aplicação dos estilos definidos para as celulas do output nas colunas alvo de acordo com
        os valores nas colunas de flag dos estilos.

        :param alvo_cols: Lista de colunas que dever receber os estilos
        :type alvo_cols: list
        :param flag_cols:
        :type flag_cols: list
        :return:
        """
        for idx, row in enumerate(self.df.iterrows(), 2):
            for boolean_col, alvo_col in product(flag_cols, alvo_cols):
                boolean_value = row[1][boolean_col]
                alvo_col_idx = self.df.columns.get_loc(alvo_col) + 1  # Pega index da coluna alvo pelo nome
                cell = self.ws_data.cell(row=idx, column=alvo_col_idx)

                if boolean_value:
                    if boolean_col == 'cinza_cma_aberto' or boolean_col == 'cinza_cma_fechado':
                        cell.style = CellStyle.CINZA_BG_STYLE.value

                    if boolean_col == 'laranja_cma_aberto' or boolean_col == 'laranja_cma_fechado':
                        cell.style = CellStyle.LARANJA_FONT_STYLE.value

                    # Caso especial caso ambas a flags estejam marcadas
                    if all(row[1][col] for col in flag_cols):
                        cell.style = CellStyle.CINZA_BG_LARANJA_FONT_STYLE.value

    def __popular_wb_dados(self):
        """
        Função responsável por popular o workbook de dados com os presentes no dataframe fornecido no
        construtor.
        """

        for r_idx, row in enumerate(dataframe_to_rows(self.df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                self.ws_data.cell(row=r_idx, column=c_idx, value=value)

    def __formatar_wb_dados(self):
        """
        Aplica as necessárias formatacões nos valores do workbook de dados
        """
        # Aplica as mudanças de cor da fonte e fundo das células de CMA
        self.style_cells_cma(alvo_cols=['CMA aberto'], flag_cols=['cinza_cma_aberto', 'laranja_cma_aberto'])
        self.style_cells_cma(alvo_cols=['CMA fechado'], flag_cols=['cinza_cma_fechado', 'laranja_cma_fechado'])

        # Centralizando texto na sheet
        for col in self.ws_data.iter_cols(min_col=1, max_col=self.ws_data.max_column):
            for r_idx, cell in enumerate(col, 1):
                alignment_obj = copy(cell.alignment)
                alignment_obj.horizontal = 'center'
                alignment_obj.vertical = 'center'
                cell.alignment = alignment_obj

        # Deixando texto em certas colunas em negrito
        bold_columns = ["Valor VOR (mg/l)", "VOR", "CMA aberto", "CMA fechado"]

        for col in bold_columns:
            col_idx = self.df.columns.get_loc(col) + 1
            for cell in self.ws_data.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for c in cell:
                    existing_color = c.font.color if c.font.color else None
                    c.font = Font(bold=True, color=existing_color)

    def __remover_elementos_desncessarios_wb_dados(self):
        """
        Remove as colunas de flag e a primeira linha da worksheet que contém os headers do dataframe
        """
        for col_idx in range(self.ws_data.max_column, self.ws_data.max_column - 4, -1):
            self.ws_data.delete_cols(col_idx)

        self.ws_data.delete_rows(1)

    def __unir_conteudo_wb_header_dados(self):
        """
        Coloca o conteúdo do worksheet de dados no worksheet de cabeçalho
        """
        last_row_source = self.ws_data.max_row

        for row in self.ws_data.iter_rows(min_row=1, max_row=last_row_source):
            new_row = []
            for cell in row:
                new_cell = self.ws_header.cell(row=self.max_linha_header + cell.row, column=cell.column, value=cell.value)

                # Copia estilo da fonte
                new_cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold,
                                     italic=cell.font.italic,
                                     color=cell.font.color)

                # Copia estilo do background
                if cell.fill:
                    new_cell.fill = PatternFill(start_color=cell.fill.start_color, end_color=cell.fill.end_color,
                                                fill_type=cell.fill.fill_type)

                # Copia alinhamento da células
                if cell.alignment:
                    new_cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        text_rotation=cell.alignment.text_rotation,
                        wrap_text=cell.alignment.wrap_text,
                        shrink_to_fit=cell.alignment.shrink_to_fit,
                        indent=cell.alignment.indent
                    )

                new_row.append(new_cell)

    def __mesclar_células_nan(self):
        """
        Realiza a mesclagem das células vazias na worksheet para manter de acordo com os padrões definidos
        para o output
        """
        for col in self.ws_header.iter_cols(min_col=1, max_col=self.ws_header.max_column):
            self.merge_cells_below_nan(col, self.max_linha_header + 1)

    def salvar_workbook_formatado(self, output_path: str):
        """
        Salva o workbook formatado no local desejado

        :param output_path:
        """
        self.wb_header.save(output_path)